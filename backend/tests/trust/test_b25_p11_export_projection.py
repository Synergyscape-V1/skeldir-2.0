"""B2.5-P11 display honesty, format validity, and injection proofs."""

from __future__ import annotations

import asyncio
import csv
import concurrent.futures
import gc
import json
import threading
import time
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO, StringIO
from pathlib import Path
from uuid import uuid4

from jsonschema import Draft202012Validator
from fastapi import FastAPI
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
import psutil
import pytest
import yaml

from app.api import export as export_api
from app.db.deps import get_db_session
from app.security.auth import get_auth_context
from app.api import export as export_module
from app.api.export import (
    COMPAT_CSV_COLUMNS,
    COMPAT_CSV_MEDIA_TYPE,
    COMPAT_CSV_SCHEMA_VERSION,
    CSV_COLUMNS,
    CSV_SCHEMA_VERSION,
    ENRICHED_CSV_MEDIA_TYPE,
    LEGACY_CSV_COLUMNS,
    LEGACY_CSV_SCHEMA_VERSION,
    NON_AUTHORITATIVE_DISPLAY,
    RETIRED_CSV_SCHEMA_VERSIONS,
    SUPPORTED_CSV_SCHEMA_VERSIONS,
    EXPORT_ROW_ALLOWLIST,
    LEGACY_EXPORT_MAX_BYTES,
    LEGACY_EXPORT_MAX_ROWS,
    LegacyExportDeadlineExceeded,
    TRACK1_MAX_CONCURRENT_EXPORTS,
    TRACK1_MAX_SERIALIZATION_WORKING_SET_BYTES,
    XLSX_COLUMNS,
    _compat_csv_from_rows,
    _csv_from_rows,
    _enforce_export_payload_no_leak,
    _xlsx_from_projection,
)
from app.trust.export_projection import (
    PROJECTION_AUTHORITY_NON_AUTHORITATIVE,
    PROJECTION_SCHEMA_VERSION,
    build_display_projection,
    project_display_rows,
)
from app.trust.spreadsheet_safety import (
    SpreadsheetSafetyError,
    neutralize_spreadsheet_cell,
)


ROOT = Path(__file__).resolve().parents[3]


def _walk(value):
    yield value
    if isinstance(value, dict):
        for child in value.values():
            yield from _walk(child)
    elif isinstance(value, list):
        for child in value:
            yield from _walk(child)


def _projection(channels: list[str] | None = None) -> tuple[object, dict[str, object]]:
    tenant_id = uuid4()
    values = channels or ["Zulu", "Alpha"]
    source_rows = [
        {
            "export_date": date(2026, 8, 8),
            "channel_code": value,
            "revenue_cents": 12_505 + index,
            "conversion_count": index + 1,
            "confidence_score": Decimal("0.92"),
        }
        for index, value in enumerate(values)
    ]
    payload = build_display_projection(
        tenant_id=tenant_id,
        start=date(2026, 8, 1),
        end=date(2026, 8, 8),
        source_rows=reversed(source_rows),
        generated_at=datetime(2026, 8, 8, 12, 0, tzinfo=timezone.utc),
    )
    return tenant_id, payload


def test_display_projection_is_closed_honest_tenant_safe_and_float_free() -> None:
    tenant_id, payload = _projection()
    schema = json.loads(
        (ROOT / "contracts/trust-api/export-projection.schema.json").read_text(
            encoding="utf-8"
        )
    )
    Draft202012Validator(schema).validate(payload)
    _enforce_export_payload_no_leak(payload)
    assert payload["projection_authority"] == PROJECTION_AUTHORITY_NON_AUTHORITATIVE
    assert payload["projection_schema_version"] == PROJECTION_SCHEMA_VERSION
    assert "tenant_id" not in payload
    assert str(tenant_id) not in json.dumps(payload, ensure_ascii=False)
    assert all(not isinstance(value, float) for value in _walk(payload))
    rows = payload["rows"]
    assert isinstance(rows, list)
    assert [row["channel"] for row in rows] == ["Alpha", "Zulu"]
    assert all(set(row) == set(EXPORT_ROW_ALLOWLIST) for row in rows)
    assert rows[0]["revenue_display"] == "125.06"
    assert rows[0]["confidence_display"] == "0.92"


def test_projection_order_is_independent_of_database_result_order() -> None:
    source = [
        {
            "export_date": date(2026, 8, 9),
            "channel_code": "Bravo",
            "revenue_cents": 2,
            "conversion_count": 1,
            "confidence_score": Decimal("0.50"),
        },
        {
            "export_date": date(2026, 8, 8),
            "channel_code": "Zulu",
            "revenue_cents": 1,
            "conversion_count": 1,
            "confidence_score": Decimal("0.50"),
        },
        {
            "export_date": date(2026, 8, 8),
            "channel_code": "Alpha",
            "revenue_cents": 3,
            "conversion_count": 1,
            "confidence_score": Decimal("0.50"),
        },
    ]
    assert project_display_rows(source) == project_display_rows(reversed(source))


def test_csv_round_trip_handles_adversarial_text_without_column_drift() -> None:
    channels = [
        "comma,value",
        'quote"value',
        "line\nvalue",
        "carriage\rvalue",
        "東京-Café",
        "",
        "x" * 256,
    ]
    _, payload = _projection(channels)
    rows = payload["rows"]
    assert isinstance(rows, list)
    rendered = _csv_from_rows(rows)
    parsed = list(csv.reader(StringIO(rendered, newline="")))
    assert tuple(parsed[0]) == CSV_COLUMNS
    assert len(parsed) == len(rows) + 1
    assert all(
        record[0] == PROJECTION_AUTHORITY_NON_AUTHORITATIVE for record in parsed[1:]
    )
    assert all(record[1] == CSV_SCHEMA_VERSION for record in parsed[1:])
    assert all(len(record) == len(CSV_COLUMNS) for record in parsed)
    assert [record[3] for record in parsed[1:]] == [str(row["channel"]) for row in rows]
    assert rendered.endswith("\r\n")


def test_compat_csv_default_is_positionally_unchanged_and_self_identifying() -> None:
    """P11-G4 and positional compatibility must hold simultaneously.

    Indices 0..4 remain the exact original five-column legacy contract, so any
    consumer reading by position is unaffected. Indices 5..6 carry the authority
    classification, so the detached file can still state that it is
    non-authoritative without a preamble.
    """
    _, payload = _projection(["Meta"])
    rows = payload["rows"]
    assert isinstance(rows, list)
    records = list(csv.reader(StringIO(_compat_csv_from_rows(rows), newline="")))

    # Positional compatibility: original five columns, original order, original values.
    assert tuple(records[0][:5]) == LEGACY_CSV_COLUMNS
    assert records[1][:5] == ["2026-08-08", "Meta", "125.05", "1", "0.92"]

    # P11-G4: the artifact identifies its own authority class.
    assert tuple(records[0]) == COMPAT_CSV_COLUMNS
    assert records[1][5] == NON_AUTHORITATIVE_DISPLAY
    assert records[1][6] == COMPAT_CSV_SCHEMA_VERSION
    assert COMPAT_CSV_SCHEMA_VERSION == "b25-p11-export-csv-compat-v1"


def test_every_active_csv_profile_satisfies_p11_g4() -> None:
    """Universal property: no active CSV profile may emit an ambiguous file.

    This is the hierarchy-aware control the second corrective cycle lacked. It
    iterates the *supported* profile set rather than a hand-picked profile, so a
    newly added profile that cannot self-identify fails here automatically.
    """
    _, payload = _projection(["Meta"])
    rows = payload["rows"]
    assert isinstance(rows, list)

    serializers = {
        COMPAT_CSV_SCHEMA_VERSION: _compat_csv_from_rows,
        CSV_SCHEMA_VERSION: _csv_from_rows,
    }
    assert set(serializers) == set(SUPPORTED_CSV_SCHEMA_VERSIONS)

    for profile, serializer in serializers.items():
        rendered = serializer(rows)
        # Detach the artifact from every HTTP affordance: only the bytes remain.
        records = list(csv.reader(StringIO(rendered, newline="")))
        header, data_rows = records[0], records[1:]
        assert data_rows, f"{profile} produced no rows"
        has_authority_column = NON_AUTHORITATIVE_DISPLAY in {
            value for record in data_rows for value in record
        }
        has_envelope_ref = "envelope_ref" in header
        assert has_authority_column or has_envelope_ref, (
            f"active CSV profile {profile} emits a detached artifact that "
            "carries neither a non-authoritative display label nor an "
            "envelope_ref, violating P11-G4"
        )


def test_retired_legacy_profile_is_not_an_active_serializer() -> None:
    """legacy-v1 must not be reachable as an emitted artifact."""
    assert LEGACY_CSV_SCHEMA_VERSION == "legacy-v1"
    assert LEGACY_CSV_SCHEMA_VERSION in RETIRED_CSV_SCHEMA_VERSIONS
    assert LEGACY_CSV_SCHEMA_VERSION not in SUPPORTED_CSV_SCHEMA_VERSIONS
    assert not hasattr(export_module, "_legacy_csv_from_rows")


def test_formula_prefixes_are_text_in_csv_and_real_xlsx() -> None:
    formulas = [
        "=1+1",
        "+SUM(A1:A2)",
        "-1+2",
        "@SUM(A1:A2)",
        '=HYPERLINK("https://invalid.example","x")',
        "\t=1+1",
        "\r=1+1",
    ]
    tenant_id, payload = _projection(formulas)
    rows = payload["rows"]
    assert isinstance(rows, list)
    assert all(str(row["channel"]).startswith("'") for row in rows)

    parsed = list(csv.reader(StringIO(_csv_from_rows(rows), newline="")))
    assert all(record[3].startswith("'") for record in parsed[1:])

    workbook_bytes = _xlsx_from_projection(payload)
    workbook = load_workbook(BytesIO(workbook_bytes), data_only=False)
    assert workbook.sheetnames == ["Export", "Metadata"]
    export_sheet = workbook["Export"]
    assert tuple(cell.value for cell in export_sheet[1]) == XLSX_COLUMNS
    for cell in export_sheet["B"][1:]:
        assert isinstance(cell.value, str) and cell.value.startswith("'")
        assert cell.data_type != "f"
    metadata_values = [cell.value for row in workbook["Metadata"] for cell in row]
    assert PROJECTION_AUTHORITY_NON_AUTHORITATIVE in metadata_values
    assert PROJECTION_SCHEMA_VERSION in metadata_values
    assert str(tenant_id) not in "|".join(str(value) for value in metadata_values)


def test_spreadsheet_transform_refuses_machine_authority_field_paths() -> None:
    assert neutralize_spreadsheet_cell("=1+1") == "'=1+1"
    try:
        neutralize_spreadsheet_cell(
            "=1+1",
            field_path="envelopes[].verified_revenue_minor",
        )
    except SpreadsheetSafetyError as exc:
        assert "machine_authority_field_neutralization_forbidden" in str(exc)
    else:
        raise AssertionError("machine-authority spreadsheet mutation was accepted")


def test_detached_csv_is_header_first_rectangular_self_identifying_at_1000_rows(
    tmp_path: Path,
) -> None:
    _, payload = _projection([f"channel-{index:04d}" for index in range(1_000)])
    rows = payload["rows"]
    assert isinstance(rows, list) and len(rows) == LEGACY_EXPORT_MAX_ROWS
    body = _csv_from_rows(rows)
    detached = tmp_path / "detached-export.csv"
    detached.write_bytes(body.encode("utf-8"))

    with detached.open("r", encoding="utf-8", newline="") as stream:
        records = list(csv.reader(stream))
    assert tuple(records[0]) == CSV_COLUMNS
    assert all(len(record) == len(CSV_COLUMNS) for record in records)
    assert {record[0] for record in records[1:]} == {
        PROJECTION_AUTHORITY_NON_AUTHORITATIVE
    }
    assert {record[1] for record in records[1:]} == {CSV_SCHEMA_VERSION}
    assert detached.stat().st_size <= LEGACY_EXPORT_MAX_BYTES
    print(
        "\nP11_TRACK1_CSV_METRICS="
        + json.dumps(
            {
                "artifact_bytes": detached.stat().st_size,
                "columns": len(CSV_COLUMNS),
                "header_first": True,
                "projection_authority": PROJECTION_AUTHORITY_NON_AUTHORITATIVE,
                "rectangular": True,
                "rows": len(records) - 1,
                "schema_version": CSV_SCHEMA_VERSION,
            },
            sort_keys=True,
        )
    )


def test_csv_and_xlsx_maximum_concurrent_serialization_stays_in_declared_memory() -> (
    None
):
    _, payload = _projection([f"channel-{index:04d}" for index in range(1_000)])
    rows = payload["rows"]
    assert isinstance(rows, list) and len(rows) == LEGACY_EXPORT_MAX_ROWS

    def serialize_pair() -> tuple[int, int]:
        csv_bytes = _csv_from_rows(rows).encode("utf-8")
        xlsx_bytes = _xlsx_from_projection(payload)
        parsed = load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=False)
        assert parsed.sheetnames == ["Export", "Metadata"]
        return len(csv_bytes), len(xlsx_bytes)

    gc.collect()
    process = psutil.Process()
    baseline_rss = process.memory_info().rss
    rss_samples = [baseline_rss]
    sampling_done = threading.Event()

    def sample_process_rss() -> None:
        while not sampling_done.wait(0.001):
            rss_samples.append(process.memory_info().rss)

    sampler = threading.Thread(target=sample_process_rss, daemon=True)
    sampler.start()
    started = time.perf_counter()
    try:
        with concurrent.futures.ThreadPoolExecutor(
            max_workers=TRACK1_MAX_CONCURRENT_EXPORTS
        ) as executor:
            sizes = list(
                executor.map(
                    lambda _: serialize_pair(), range(TRACK1_MAX_CONCURRENT_EXPORTS)
                )
            )
    finally:
        sampling_done.set()
        sampler.join(timeout=1)
        rss_samples.append(process.memory_info().rss)
    elapsed_ms = (time.perf_counter() - started) * 1_000
    peak_rss = max(rss_samples)
    peak_rss_delta = peak_rss - baseline_rss

    assert all(
        csv_bytes <= LEGACY_EXPORT_MAX_BYTES and xlsx_bytes <= LEGACY_EXPORT_MAX_BYTES
        for csv_bytes, xlsx_bytes in sizes
    )
    assert peak_rss_delta <= (
        TRACK1_MAX_CONCURRENT_EXPORTS * TRACK1_MAX_SERIALIZATION_WORKING_SET_BYTES
    )
    print(
        "\nP11_TRACK1_SERIALIZATION_METRICS="
        + json.dumps(
            {
                "concurrency": TRACK1_MAX_CONCURRENT_EXPORTS,
                "declared_aggregate_working_set_bytes": (
                    TRACK1_MAX_CONCURRENT_EXPORTS
                    * TRACK1_MAX_SERIALIZATION_WORKING_SET_BYTES
                ),
                "elapsed_ms": elapsed_ms,
                "measurement": "psutil.Process.memory_info().rss",
                "baseline_rss_bytes": baseline_rss,
                "peak_rss_bytes": peak_rss,
                "peak_rss_delta_bytes": peak_rss_delta,
                "rows_per_export": len(rows),
                "sizes": [
                    {"csv_bytes": csv_bytes, "xlsx_bytes": xlsx_bytes}
                    for csv_bytes, xlsx_bytes in sizes
                ],
            },
            sort_keys=True,
        )
    )


@pytest.mark.asyncio
async def test_track1_deadline_includes_saturated_admission_queue(monkeypatch) -> None:
    semaphore = asyncio.Semaphore(TRACK1_MAX_CONCURRENT_EXPORTS)
    for _ in range(TRACK1_MAX_CONCURRENT_EXPORTS):
        await semaphore.acquire()
    monkeypatch.setattr(export_api, "_TRACK1_EXPORT_CONCURRENCY_LIMIT", semaphore)
    monkeypatch.setattr(export_api, "TRACK1_HANDLER_DEADLINE_SECONDS", 0.02)

    projection_called = False

    async def forbidden_projection(**_kwargs):
        nonlocal projection_called
        projection_called = True
        raise AssertionError("saturated request reached database work")

    monkeypatch.setattr(export_api, "_projection_for_request", forbidden_projection)
    try:
        with pytest.raises(
            LegacyExportDeadlineExceeded,
            match="legacy_export_handler_deadline_exceeded",
        ):
            await export_api._run_track1_export(
                db_session=object(),
                tenant_id=uuid4(),
                session_scope=None,
                start=date(2026, 8, 1),
                end=date(2026, 8, 31),
                channels=None,
                export_format="csv",
            )
    finally:
        for _ in range(TRACK1_MAX_CONCURRENT_EXPORTS):
            semaphore.release()
    assert projection_called is False


@pytest.mark.asyncio
async def test_timeout_retains_capacity_until_physical_serializer_finishes(
    monkeypatch,
) -> None:
    semaphore = asyncio.Semaphore(TRACK1_MAX_CONCURRENT_EXPORTS)
    monkeypatch.setattr(export_api, "_TRACK1_EXPORT_CONCURRENCY_LIMIT", semaphore)
    monkeypatch.setattr(export_api, "TRACK1_HANDLER_DEADLINE_SECONDS", 0.03)
    export_api._TRACK1_RETAINED_SERIALIZER_TASKS.clear()
    _, payload = _projection(["Meta"])
    rows = payload["rows"]
    assert isinstance(rows, list)

    async def fast_projection(**_kwargs):
        return rows, payload

    real_serializer = _csv_from_rows
    release_physical_work = threading.Event()
    state_lock = threading.Lock()
    active = 0
    peak_active = 0
    starts = 0

    def blocked_production_serializer(value):
        nonlocal active, peak_active, starts
        with state_lock:
            starts += 1
            active += 1
            peak_active = max(peak_active, active)
        try:
            release_physical_work.wait(timeout=2)
            return real_serializer(value)
        finally:
            with state_lock:
                active -= 1

    monkeypatch.setattr(export_api, "_projection_for_request", fast_projection)
    monkeypatch.setattr(export_api, "_csv_from_rows", blocked_production_serializer)

    async def invoke() -> None:
        with pytest.raises(
            LegacyExportDeadlineExceeded,
            match="legacy_export_handler_deadline_exceeded",
        ):
            await export_api._run_track1_export(
                db_session=object(),
                tenant_id=uuid4(),
                session_scope=None,
                start=date(2026, 8, 1),
                end=date(2026, 8, 1),
                channels=None,
                export_format="csv",
                csv_schema_version=CSV_SCHEMA_VERSION,
            )

    await asyncio.gather(invoke(), invoke())
    assert active == peak_active == starts == TRACK1_MAX_CONCURRENT_EXPORTS
    assert semaphore._value == 0
    assert len(export_api._TRACK1_RETAINED_SERIALIZER_TASKS) == 2

    await invoke()
    assert starts == TRACK1_MAX_CONCURRENT_EXPORTS
    assert active == TRACK1_MAX_CONCURRENT_EXPORTS
    assert semaphore._value == 0

    release_physical_work.set()
    for _ in range(100):
        if active == 0 and semaphore._value == TRACK1_MAX_CONCURRENT_EXPORTS:
            break
        await asyncio.sleep(0.01)
    assert active == 0
    assert semaphore._value == TRACK1_MAX_CONCURRENT_EXPORTS
    assert not export_api._TRACK1_RETAINED_SERIALIZER_TASKS
    print(
        "\nP11_TRACK1_TIMEOUT_SERIALIZER_METRICS="
        + json.dumps(
            {
                "active_after_http_timeout": TRACK1_MAX_CONCURRENT_EXPORTS,
                "capacity_after_http_timeout": 0,
                "capacity_after_physical_completion": semaphore._value,
                "peak_physical_serializers": peak_active,
                "third_serializer_started_while_saturated": False,
            },
            sort_keys=True,
        )
    )


@pytest.mark.asyncio
@pytest.mark.parametrize(
    "reason_code",
    [
        "legacy_export_database_deadline_exceeded",
        "legacy_export_handler_deadline_exceeded",
    ],
)
async def test_runtime_503_matches_authoritative_openapi_schema(
    monkeypatch, reason_code: str
) -> None:
    app = FastAPI()
    app.include_router(export_api.router, prefix="/api/export")

    async def fake_auth():
        return type("Auth", (), {"tenant_id": uuid4()})()

    async def fake_db():
        yield object()

    async def deadline(**_kwargs):
        raise LegacyExportDeadlineExceeded(reason_code)

    app.dependency_overrides[get_auth_context] = fake_auth
    app.dependency_overrides[get_db_session] = fake_db
    monkeypatch.setattr(export_api, "_run_track1_export", deadline)
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as client:
        response = await client.get(
            "/api/export/csv",
            headers={"X-Correlation-ID": str(uuid4())},
        )

    assert response.status_code == 503
    contract = yaml.safe_load(
        (ROOT / "api-contracts/openapi/v1/export.yaml").read_text(encoding="utf-8")
    )
    schema = contract["components"]["schemas"]["ExportDeadlineError"]
    Draft202012Validator(schema).validate(response.json())
    assert response.json() == {
        "detail": {"status": "refused", "reason_code": reason_code}
    }


# ---------------------------------------------------------------------------
# Third corrective: runtime / contract parity for every governed refusal, and
# detached-artifact authority for the DEFAULT (no-parameter) request.
# ---------------------------------------------------------------------------

_TRACK1_CSV_ROUTES = ("/api/export/csv", "/api/export/revenue?format=csv")

_GOVERNED_413_REASONS = (
    "legacy_export_date_span_exceeded",
    "legacy_export_channel_count_exceeded",
    "legacy_export_channel_length_exceeded",
    "legacy_export_row_admission_exceeded",
    "legacy_export_byte_admission_exceeded",
    "legacy_export_row_budget_exceeded",
    "legacy_export_byte_budget_exceeded",
)


def _export_app() -> FastAPI:
    app = FastAPI()
    app.include_router(export_api.router, prefix="/api/export")

    async def fake_auth():
        return type("Auth", (), {"tenant_id": uuid4()})()

    async def fake_db():
        yield object()

    app.dependency_overrides[get_auth_context] = fake_auth
    app.dependency_overrides[get_db_session] = fake_db
    return app


def _canonical_contract() -> dict:
    return yaml.safe_load(
        (ROOT / "api-contracts/openapi/v1/export.yaml").read_text(encoding="utf-8")
    )


@pytest.mark.asyncio
@pytest.mark.parametrize("reason_code", _GOVERNED_413_REASONS)
async def test_runtime_413_matches_authoritative_openapi_schema(
    monkeypatch, reason_code: str
) -> None:
    """Gate P11-C3-I: every deliberate 413 is representable in the contract."""
    app = _export_app()

    async def limit(**_kwargs):
        raise export_api.LegacyExportLimitExceeded(reason_code)

    monkeypatch.setattr(export_api, "_run_track1_export", limit)
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as client:
        response = await client.get(
            "/api/export/csv", headers={"X-Correlation-ID": str(uuid4())}
        )

    assert response.status_code == 413
    schema = _canonical_contract()["components"]["schemas"]["ExportLimitError"]
    Draft202012Validator(schema).validate(response.json())
    assert response.json() == {
        "detail": {"status": "refused", "reason_code": reason_code}
    }


@pytest.mark.asyncio
@pytest.mark.parametrize("route", _TRACK1_CSV_ROUTES)
async def test_retired_csv_profile_is_refused_with_governed_410(route: str) -> None:
    """Gate P11-C3-A/B: the ambiguous profile is refused, never emitted."""
    app = _export_app()
    separator = "&" if "?" in route else "?"
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as client:
        response = await client.get(
            f"{route}{separator}csv_schema_version=legacy-v1",
            headers={"X-Correlation-ID": str(uuid4())},
        )

    assert response.status_code == 410
    schema = _canonical_contract()["components"]["schemas"]["ExportProfileRetiredError"]
    Draft202012Validator(schema).validate(response.json())
    detail = response.json()["detail"]
    assert detail["reason_code"] == "legacy_csv_profile_retired"
    assert detail["retired_profile"] == LEGACY_CSV_SCHEMA_VERSION
    assert detail["replacement_profile"] == COMPAT_CSV_SCHEMA_VERSION


@pytest.mark.asyncio
@pytest.mark.parametrize("route", _TRACK1_CSV_ROUTES)
async def test_default_request_emits_self_identifying_detached_csv(
    monkeypatch, route: str
) -> None:
    """Gate P11-C3-A: the DEFAULT artifact, with the HTTP context discarded.

    The request deliberately omits `csv_schema_version`. Only the response body
    is retained -- no URL, no query parameters, no headers, no media type -- and
    the file must still declare that it is non-authoritative.
    """
    app = _export_app()
    _, payload = _projection(["Meta"])
    rows = payload["rows"]

    async def fake_export(*, export_format, csv_schema_version, **_kwargs):
        assert (
            csv_schema_version == COMPAT_CSV_SCHEMA_VERSION
        ), "default profile drifted away from the compliant compat profile"
        return rows, payload, _compat_csv_from_rows(rows)

    monkeypatch.setattr(export_api, "_run_track1_export", fake_export)
    async with AsyncClient(
        transport=ASGITransport(app=app), base_url="http://test"
    ) as client:
        response = await client.get(route, headers={"X-Correlation-ID": str(uuid4())})

    assert response.status_code == 200

    # Everything except the bytes is now discarded.
    detached_bytes = response.content
    records = list(csv.reader(StringIO(detached_bytes.decode("utf-8"), newline="")))
    header, data_rows = records[0], [r for r in records[1:] if r]

    assert tuple(header[:5]) == LEGACY_CSV_COLUMNS
    assert "projection_authority" in header
    assert data_rows
    for record in data_rows:
        assert (
            NON_AUTHORITATIVE_DISPLAY in record
        ), "default detached CSV cannot identify itself as non-authoritative"


@pytest.mark.asyncio
@pytest.mark.parametrize("route", _TRACK1_CSV_ROUTES)
async def test_csv_media_type_profile_matches_emitted_profile(
    monkeypatch, route: str
) -> None:
    """Gate P11-C3-J: media type must name the profile actually emitted."""
    _, payload = _projection(["Meta"])
    rows = payload["rows"]

    expected = {
        COMPAT_CSV_SCHEMA_VERSION: (COMPAT_CSV_MEDIA_TYPE, _compat_csv_from_rows),
        CSV_SCHEMA_VERSION: (ENRICHED_CSV_MEDIA_TYPE, _csv_from_rows),
    }
    assert set(expected) == set(SUPPORTED_CSV_SCHEMA_VERSIONS)

    for profile, (media_type, serializer) in expected.items():
        app = _export_app()

        async def fake_export(*, export_format, csv_schema_version, **_kwargs):
            return rows, payload, serializer(rows)

        monkeypatch.setattr(export_api, "_run_track1_export", fake_export)
        separator = "&" if "?" in route else "?"
        async with AsyncClient(
            transport=ASGITransport(app=app), base_url="http://test"
        ) as client:
            response = await client.get(
                f"{route}{separator}csv_schema_version={profile}",
                headers={"X-Correlation-ID": str(uuid4())},
            )

        assert response.status_code == 200
        assert response.headers["content-type"].startswith(media_type), (
            f"profile {profile} emitted content-type "
            f"{response.headers['content-type']!r}, expected {media_type!r}"
        )
