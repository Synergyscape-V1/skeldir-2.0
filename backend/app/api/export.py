"""Human-authorized, non-authoritative compatibility export routes."""

from __future__ import annotations

import asyncio
import csv
import logging
from collections.abc import Callable
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO, StringIO
from typing import Annotated, Any, Literal
from uuid import UUID

from fastapi import (
    APIRouter,
    Depends,
    Header,
    HTTPException,
    Query,
    Response,
    Security,
    status,
)
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.exc import DBAPIError
from sqlalchemy.ext.asyncio import AsyncSession

from app.db.deps import get_db_session
from app.privacy.output_redaction import (
    find_output_leaks,
    output_forbidden_key_set,
    sanitize_output_payload,
)
from app.security.auth import AuthContext, get_auth_context
from app.trust.export_projection import build_display_projection, project_display_rows
from app.trust.refusal import tenant_hash


router = APIRouter()
logger = logging.getLogger(__name__)

EXPORT_ROW_ALLOWLIST = (
    "date",
    "channel",
    "revenue_minor",
    "revenue_display",
    "conversions",
    "confidence_display",
)
EXPORT_TOP_LEVEL_ALLOWLIST = (
    "projection_authority",
    "projection_schema_version",
    "tenant_id_hash",
    "generated_at",
    "date_range",
    "rows",
)
EXPORT_DATE_RANGE_ALLOWLIST = ("start", "end")
CSV_SCHEMA_VERSION = "b25-p11-export-csv-v2"
#: Default profile. Preserves the five legacy columns at their exact positional
#: indices 0..4 and appends the P11-G4 authority classification as trailing
#: columns 5..6. Index-based legacy readers are unaffected; a detached file can
#: still identify itself as non-authoritative. See contracts/export/CSV_EVOLUTION.md.
COMPAT_CSV_SCHEMA_VERSION = "b25-p11-export-csv-compat-v1"
#: Retired. The bare five-column shape cannot carry an authority classification
#: without either breaking positional compatibility or adding a forbidden
#: preamble, so it cannot satisfy P11-G4 and is refused rather than emitted.
LEGACY_CSV_SCHEMA_VERSION = "legacy-v1"
ENRICHED_CSV_MEDIA_TYPE = (
    'text/csv; profile="https://api.skeldir.com/profiles/export-csv-v2"'
)
COMPAT_CSV_MEDIA_TYPE = (
    'text/csv; profile="https://api.skeldir.com/profiles/export-csv-compat-v1"'
)
LEGACY_CSV_COLUMNS = ("date", "channel", "revenue", "conversions", "confidence")
#: Legacy positions 0..4 preserved byte-for-byte, authority appended at 5..6.
COMPAT_CSV_COLUMNS = LEGACY_CSV_COLUMNS + (
    "projection_authority",
    "projection_schema_version",
)
CSV_COLUMNS = (
    "projection_authority",
    "projection_schema_version",
    "date",
    "channel",
    "revenue",
    "conversions",
    "confidence",
)
NON_AUTHORITATIVE_DISPLAY = "non_authoritative_display"
SUPPORTED_CSV_SCHEMA_VERSIONS = (COMPAT_CSV_SCHEMA_VERSION, CSV_SCHEMA_VERSION)
RETIRED_CSV_SCHEMA_VERSIONS = (LEGACY_CSV_SCHEMA_VERSION,)
XLSX_COLUMNS = ("date", "channel", "revenue", "conversions", "confidence")

LEGACY_EXPORT_MAX_DATE_SPAN_DAYS = 31
LEGACY_EXPORT_MAX_ROWS = 1_000
LEGACY_EXPORT_MAX_BYTES = 1_048_576
LEGACY_EXPORT_MAX_CHANNELS = 32
LEGACY_EXPORT_MAX_CHANNEL_LENGTH = 256
TRACK1_MAX_CONCURRENT_EXPORTS = 2
TRACK1_HANDLER_DEADLINE_SECONDS = 3.0
TRACK1_DATABASE_STATEMENT_TIMEOUT_MS = 1_250
TRACK1_DATABASE_WORK_MEM_KIB = 4_096
TRACK1_MAX_SERIALIZATION_WORKING_SET_BYTES = 32 * 1_024 * 1_024
_LEGACY_ROW_BYTE_ESTIMATE = 512
_EXPORT_FORBIDDEN_KEYS = output_forbidden_key_set()
_TRACK1_EXPORT_CONCURRENCY_LIMIT = asyncio.Semaphore(TRACK1_MAX_CONCURRENT_EXPORTS)
_TRACK1_RETAINED_SERIALIZER_TASKS: set[asyncio.Task[Any]] = set()


class LegacyExportLimitExceeded(ValueError):
    """Raised before unbounded legacy export work can occur."""


class LegacyExportDeadlineExceeded(TimeoutError):
    """Raised when a bounded Track-1 database or handler deadline fires."""


def _resolve_date_range(
    *,
    start_date: date | None,
    end_date: date | None,
) -> tuple[date, date]:
    today = datetime.now(timezone.utc).date()
    start = start_date or today
    end = end_date or start
    if start > end:
        start, end = end, start
    span_days = (end - start).days + 1
    if span_days > LEGACY_EXPORT_MAX_DATE_SPAN_DAYS:
        raise LegacyExportLimitExceeded("legacy_export_date_span_exceeded")
    return start, end


def _admit_legacy_export(
    *,
    start: date,
    end: date,
    channels: list[str] | None,
) -> list[str]:
    """Reject impossible row/byte envelopes before opening the SQL result."""
    normalized = [
        value.strip().lower() for value in (channels or []) if value and value.strip()
    ]
    if len(normalized) > LEGACY_EXPORT_MAX_CHANNELS:
        raise LegacyExportLimitExceeded("legacy_export_channel_count_exceeded")
    if any(len(value) > LEGACY_EXPORT_MAX_CHANNEL_LENGTH for value in normalized):
        raise LegacyExportLimitExceeded("legacy_export_channel_length_exceeded")
    span_days = (end - start).days + 1
    admitted_channel_count = len(normalized) or LEGACY_EXPORT_MAX_CHANNELS
    admitted_rows = span_days * admitted_channel_count
    if admitted_rows > LEGACY_EXPORT_MAX_ROWS:
        raise LegacyExportLimitExceeded("legacy_export_row_admission_exceeded")
    if admitted_rows * _LEGACY_ROW_BYTE_ESTIMATE > LEGACY_EXPORT_MAX_BYTES:
        raise LegacyExportLimitExceeded("legacy_export_byte_admission_exceeded")
    return normalized


def _csv_from_rows(rows: list[dict[str, object]]) -> str:
    stream = StringIO(newline="")
    writer = csv.writer(
        stream, dialect="excel", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n"
    )
    writer.writerow(CSV_COLUMNS)
    for row in rows:
        _enforce_export_row_no_leak(row)
        writer.writerow(
            (
                "non_authoritative_display",
                CSV_SCHEMA_VERSION,
                row["date"],
                row["channel"],
                row["revenue_display"],
                row["conversions"],
                row["confidence_display"],
            )
        )
    rendered = stream.getvalue()
    if len(rendered.encode("utf-8")) > LEGACY_EXPORT_MAX_BYTES:
        raise LegacyExportLimitExceeded("legacy_export_byte_budget_exceeded")
    return rendered


def _compat_csv_from_rows(rows: list[dict[str, object]]) -> str:
    """Emit the positional-compatible, self-identifying default CSV profile.

    Columns 0..4 are the original five legacy columns in their exact original
    order, so any consumer reading by positional index is unaffected. Columns
    5..6 carry the P11-G4 authority classification so the detached file can
    still identify itself as non-authoritative without a preamble.
    """
    stream = StringIO(newline="")
    writer = csv.writer(
        stream, dialect="excel", quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n"
    )
    writer.writerow(COMPAT_CSV_COLUMNS)
    for row in rows:
        _enforce_export_row_no_leak(row)
        writer.writerow(
            (
                row["date"],
                row["channel"],
                row["revenue_display"],
                row["conversions"],
                row["confidence_display"],
                NON_AUTHORITATIVE_DISPLAY,
                COMPAT_CSV_SCHEMA_VERSION,
            )
        )
    rendered = stream.getvalue()
    if len(rendered.encode("utf-8")) > LEGACY_EXPORT_MAX_BYTES:
        raise LegacyExportLimitExceeded("legacy_export_byte_budget_exceeded")
    return rendered


def _xlsx_from_projection(payload: dict[str, object]) -> bytes:
    _enforce_export_payload_no_leak(payload)
    workbook = Workbook(write_only=True)
    worksheet = workbook.create_sheet("Export")
    worksheet.append(XLSX_COLUMNS)
    rows = payload["rows"]
    if not isinstance(rows, list):
        raise ValueError("export payload rows must be a list")
    for row in rows:
        if not isinstance(row, dict):
            raise ValueError("export payload row must be an object")
        _enforce_export_row_no_leak(row)
        worksheet.append(
            (
                row["date"],
                row["channel"],
                row["revenue_display"],
                row["conversions"],
                row["confidence_display"],
            )
        )

    metadata = workbook.create_sheet("Metadata")
    metadata.append(("projection_authority", payload["projection_authority"]))
    metadata.append(("projection_schema_version", payload["projection_schema_version"]))
    metadata.append(("tenant_id_hash", payload["tenant_id_hash"]))
    metadata.append(("generated_at", payload["generated_at"]))
    date_range = payload["date_range"]
    if not isinstance(date_range, dict):
        raise ValueError("export payload date_range must be an object")
    metadata.append(("date_range_start", date_range["start"]))
    metadata.append(("date_range_end", date_range["end"]))

    stream = BytesIO()
    workbook.save(stream)
    body = stream.getvalue()
    if len(body) > LEGACY_EXPORT_MAX_BYTES:
        raise LegacyExportLimitExceeded("legacy_export_byte_budget_exceeded")
    return body


def _export_payload(
    *,
    tenant_id: UUID,
    start: date,
    end: date,
    rows: list[dict[str, object]],
) -> dict[str, object]:
    payload = build_display_projection(
        tenant_id=tenant_id,
        start=start,
        end=end,
        source_rows=rows,
    )
    _enforce_export_payload_no_leak(payload)
    sanitized = sanitize_output_payload(payload, forbidden_keys=_EXPORT_FORBIDDEN_KEYS)
    if not isinstance(sanitized, dict):
        raise ValueError("sanitized export payload must remain an object")
    return sanitized


def _filter_export_row_fields(row: dict[str, object]) -> dict[str, object]:
    return {key: row[key] for key in EXPORT_ROW_ALLOWLIST}


def _enforce_export_row_no_leak(row: dict[str, object]) -> None:
    if set(row.keys()) != set(EXPORT_ROW_ALLOWLIST):
        raise ValueError("export row contains non-allowlisted keys")
    leaks = find_output_leaks(row, forbidden_keys=_EXPORT_FORBIDDEN_KEYS)
    if leaks:
        raise ValueError(f"export row violates no-leak policy at {', '.join(leaks)}")


def _enforce_export_payload_no_leak(payload: dict[str, object]) -> None:
    if set(payload.keys()) != set(EXPORT_TOP_LEVEL_ALLOWLIST):
        raise ValueError("export payload contains non-allowlisted top-level keys")
    if payload.get("projection_authority") != "non_authoritative_display":
        raise ValueError("export projection authority marker missing")
    date_range = payload.get("date_range")
    if not isinstance(date_range, dict) or set(date_range.keys()) != set(
        EXPORT_DATE_RANGE_ALLOWLIST
    ):
        raise ValueError("export payload date_range contains non-allowlisted keys")
    rows = payload.get("rows")
    if not isinstance(rows, list):
        raise ValueError("export payload rows must be a list")
    for row in rows:
        if not isinstance(row, dict):
            raise ValueError("export payload row must be an object")
        _enforce_export_row_no_leak(row)
    leaks = find_output_leaks(payload, forbidden_keys=_EXPORT_FORBIDDEN_KEYS)
    if leaks:
        raise ValueError(
            f"export payload violates no-leak policy at {', '.join(leaks)}"
        )


async def _fetch_reporting_rows(
    *,
    db_session: AsyncSession,
    tenant_id: UUID,
    session_scope: UUID | None,
    start: date,
    end: date,
    channels: list[str] | None = None,
) -> list[dict[str, object]]:
    channel_filter = _admit_legacy_export(start=start, end=end, channels=channels)
    start_ts = datetime.combine(start, time.min, tzinfo=timezone.utc)
    end_ts = datetime.combine(end + timedelta(days=1), time.min, tzinfo=timezone.utc)
    query_params: dict[str, object] = {
        "tenant_id": str(tenant_id),
        "start_ts": start_ts,
        "end_ts": end_ts,
        "channels_is_empty": len(channel_filter) == 0,
        "channels": channel_filter,
        "row_limit": LEGACY_EXPORT_MAX_ROWS + 1,
    }
    if session_scope is None:
        session_join = ""
        session_predicate = ""
    else:
        session_join = """
            JOIN session_authority sa
              ON sa.tenant_id = e.tenant_id
             AND sa.session_id = e.session_id
        """
        session_predicate = """
              AND e.session_id = :session_id
              AND sa.invalidated_at IS NULL
              AND sa.expires_at > :authority_now
        """
        query_params.update(
            {
                "session_id": str(session_scope),
                "authority_now": datetime.now(timezone.utc),
            }
        )
    query = text(
        f"""
        SELECT
            date_trunc('day', e.occurred_at)::date AS export_date,
            aa.channel_code AS channel_code,
            COALESCE(SUM(aa.allocated_revenue_cents), 0)::bigint AS revenue_cents,
            COUNT(DISTINCT aa.event_id)::bigint AS conversion_count,
            COALESCE(AVG(aa.confidence_score), 0)::numeric AS confidence_score
        FROM attribution_allocations aa
        JOIN attribution_events e
          ON e.id = aa.event_id
         AND e.tenant_id = aa.tenant_id
        {session_join}
        WHERE aa.tenant_id = :tenant_id
          AND e.occurred_at >= :start_ts
          AND e.occurred_at < :end_ts
          {session_predicate}
          AND (
                :channels_is_empty
                OR lower(aa.channel_code) = ANY(CAST(:channels AS text[]))
          )
        GROUP BY export_date, aa.channel_code
        ORDER BY export_date ASC, aa.channel_code ASC
        LIMIT :row_limit
        """
    )
    await db_session.execute(
        text(
            f"SET LOCAL statement_timeout = '{TRACK1_DATABASE_STATEMENT_TIMEOUT_MS}ms'"
        )
    )
    await db_session.execute(
        text(f"SET LOCAL work_mem = '{TRACK1_DATABASE_WORK_MEM_KIB}kB'")
    )
    await db_session.execute(text("SET LOCAL max_parallel_workers_per_gather = 0"))
    try:
        result = await db_session.execute(query, query_params)
    except DBAPIError as exc:
        sqlstate = getattr(exc.orig, "sqlstate", None) or getattr(
            exc.orig, "pgcode", None
        )
        if sqlstate == "57014":
            raise LegacyExportDeadlineExceeded(
                "legacy_export_database_deadline_exceeded"
            ) from exc
        raise
    fetched = result.fetchmany(LEGACY_EXPORT_MAX_ROWS + 1)
    if len(fetched) > LEGACY_EXPORT_MAX_ROWS:
        raise LegacyExportLimitExceeded("legacy_export_row_budget_exceeded")
    source_rows = [
        {
            "export_date": export_date,
            "channel_code": str(channel_code),
            "revenue_cents": int(revenue_cents),
            "conversion_count": int(conversion_count),
            "confidence_score": confidence_score,
        }
        for export_date, channel_code, revenue_cents, conversion_count, confidence_score in fetched
    ]
    rows = [_filter_export_row_fields(row) for row in project_display_rows(source_rows)]
    for row in rows:
        _enforce_export_row_no_leak(row)
    return rows


def _limit_error(exc: LegacyExportLimitExceeded) -> HTTPException:
    return HTTPException(
        status_code=status.HTTP_413_CONTENT_TOO_LARGE,
        detail={"status": "refused", "reason_code": str(exc)},
    )


def _assert_csv_profile_supported(csv_schema_version: str) -> None:
    """Refuse retired CSV profiles before any tenant work is admitted.

    ``legacy-v1`` emitted a bare five-column artifact that could carry neither
    ``envelope_ref`` nor a non-authoritative display label, so a detached file
    could not state its own authority class. That is a P11-G4 violation which no
    amount of documentation can cure, so the profile is retired rather than
    silently emitted. The replacement default,
    ``b25-p11-export-csv-compat-v1``, keeps the same five columns at the same
    positional indices and appends the authority classification.
    """
    if csv_schema_version in RETIRED_CSV_SCHEMA_VERSIONS:
        raise HTTPException(
            status_code=status.HTTP_410_GONE,
            detail={
                "status": "refused",
                "reason_code": "legacy_csv_profile_retired",
                "retired_profile": csv_schema_version,
                "replacement_profile": COMPAT_CSV_SCHEMA_VERSION,
            },
        )


async def _projection_for_request(
    *,
    db_session: AsyncSession,
    tenant_id: UUID,
    session_scope: UUID | None,
    start: date,
    end: date,
    channels: list[str] | None = None,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    rows = await _fetch_reporting_rows(
        db_session=db_session,
        tenant_id=tenant_id,
        session_scope=session_scope,
        start=start,
        end=end,
        channels=channels,
    )
    payload = _export_payload(tenant_id=tenant_id, start=start, end=end, rows=rows)
    return rows, payload


async def _run_track1_export(
    *,
    db_session: AsyncSession,
    tenant_id: UUID,
    session_scope: UUID | None,
    start: date,
    end: date,
    channels: list[str] | None,
    export_format: str,
    csv_schema_version: str = COMPAT_CSV_SCHEMA_VERSION,
) -> tuple[list[dict[str, object]], dict[str, object], str | bytes | None]:
    """Bound DB and physical serializer lifetime, including after cancellation."""
    permit_acquired = False
    serializer_task: asyncio.Task[str | bytes] | None = None

    def release_retained_permit(completed: asyncio.Task[str | bytes]) -> None:
        # Retrieve executor exceptions so a timed-out HTTP request cannot create
        # an unobserved task failure. The physical permit is released only when
        # the underlying to_thread work has actually returned.
        try:
            completed.exception()
        except (asyncio.CancelledError, Exception):
            pass
        _TRACK1_RETAINED_SERIALIZER_TASKS.discard(completed)
        _TRACK1_EXPORT_CONCURRENCY_LIMIT.release()

    try:
        # The deadline includes admission queueing as well as database and
        # serialization work; saturation cannot create an unbounded wait.
        async with asyncio.timeout(TRACK1_HANDLER_DEADLINE_SECONDS):
            await _TRACK1_EXPORT_CONCURRENCY_LIMIT.acquire()
            permit_acquired = True
            rows, payload = await _projection_for_request(
                db_session=db_session,
                tenant_id=tenant_id,
                session_scope=session_scope,
                start=start,
                end=end,
                channels=channels,
            )
            serializer: Callable[[Any], str | bytes] | None
            serializer_input: Any
            if export_format == "csv":
                if csv_schema_version == CSV_SCHEMA_VERSION:
                    serializer = _csv_from_rows
                elif csv_schema_version == COMPAT_CSV_SCHEMA_VERSION:
                    serializer = _compat_csv_from_rows
                else:
                    # Retired profiles are refused at the route boundary before
                    # any work is admitted; reaching here is an invariant break.
                    raise ValueError("unsupported_csv_schema_version")
                serializer_input = rows
            elif export_format == "xlsx":
                serializer = _xlsx_from_projection
                serializer_input = payload
            elif export_format == "json":
                serializer = None
                serializer_input = None
            else:
                raise ValueError("unsupported_track1_export_format")

            if serializer is None:
                body: str | bytes | None = None
            else:
                serializer_task = asyncio.create_task(
                    asyncio.to_thread(serializer, serializer_input)
                )
                # asyncio cancellation cannot stop an executor thread. Shield
                # it, and retain the physical permit until its task completes.
                body = await asyncio.shield(serializer_task)
            return rows, payload, body
    except LegacyExportDeadlineExceeded:
        raise
    except TimeoutError as exc:
        raise LegacyExportDeadlineExceeded(
            "legacy_export_handler_deadline_exceeded"
        ) from exc
    finally:
        if permit_acquired:
            if serializer_task is not None and not serializer_task.done():
                _TRACK1_RETAINED_SERIALIZER_TASKS.add(serializer_task)
                serializer_task.add_done_callback(release_retained_permit)
            else:
                _TRACK1_EXPORT_CONCURRENCY_LIMIT.release()


def _observe_human_export(
    *,
    tenant_id: UUID,
    correlation_id: UUID,
    export_format: str,
    row_count: int,
    artifact_bytes: int,
) -> None:
    """Use ordinary application observability for human display downloads."""
    logger.info(
        "human_non_authoritative_export_completed",
        extra={
            "tenant_id_hash": tenant_hash(tenant_id),
            "correlation_id": str(correlation_id),
            "export_format": export_format,
            "projection_authority": "non_authoritative_display",
            "row_count": row_count,
            "artifact_bytes": artifact_bytes,
        },
    )


def _deadline_error(exc: LegacyExportDeadlineExceeded) -> HTTPException:
    return HTTPException(
        status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
        detail={"status": "refused", "reason_code": str(exc)},
    )


@router.get("/revenue", operation_id="exportRevenue")
async def export_revenue(
    response: Response,
    x_correlation_id: Annotated[UUID, Header(alias="X-Correlation-ID")],
    auth_context: Annotated[AuthContext, Security(get_auth_context, scopes=["viewer"])],
    db_session: Annotated[AsyncSession, Depends(get_db_session)],
    export_format: Annotated[str, Query(alias="format")] = "csv",
    start_date: date | None = Query(default=None),
    end_date: date | None = Query(default=None),
    channels: list[str] | None = Query(default=None),
    csv_schema_version: Literal[
        "b25-p11-export-csv-compat-v1", "b25-p11-export-csv-v2", "legacy-v1"
    ] = Query(default=COMPAT_CSV_SCHEMA_VERSION),
    x_attribution_session_id: Annotated[
        UUID | None, Header(alias="X-Attribution-Session-ID")
    ] = None,
):
    response.headers["X-Correlation-ID"] = str(x_correlation_id)
    try:
        start, end = _resolve_date_range(start_date=start_date, end_date=end_date)
        normalized_format = export_format.strip().lower()
        if normalized_format not in {"csv", "xlsx", "json"}:
            raise HTTPException(status_code=400, detail="Unsupported export format.")
        if normalized_format == "csv":
            _assert_csv_profile_supported(csv_schema_version)
        rows, payload, serialized = await _run_track1_export(
            db_session=db_session,
            tenant_id=auth_context.tenant_id,
            session_scope=x_attribution_session_id,
            start=start,
            end=end,
            channels=channels,
            export_format=normalized_format,
            csv_schema_version=csv_schema_version,
        )
        if normalized_format == "csv":
            if not isinstance(serialized, str):
                raise ValueError("csv_serialization_missing")
            body: str | bytes = serialized
            media_type = (
                ENRICHED_CSV_MEDIA_TYPE
                if csv_schema_version == CSV_SCHEMA_VERSION
                else COMPAT_CSV_MEDIA_TYPE
            )
            filename = "skeldir-export-revenue.csv"
        elif normalized_format == "xlsx":
            if not isinstance(serialized, bytes):
                raise ValueError("xlsx_serialization_missing")
            body = serialized
            media_type = (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            filename = "skeldir-export-revenue.xlsx"
        elif normalized_format == "json":
            _observe_human_export(
                tenant_id=auth_context.tenant_id,
                correlation_id=x_correlation_id,
                export_format=normalized_format,
                row_count=len(rows),
                artifact_bytes=len(str(payload).encode("utf-8")),
            )
            return payload
    except LegacyExportLimitExceeded as exc:
        raise _limit_error(exc) from exc
    except LegacyExportDeadlineExceeded as exc:
        raise _deadline_error(exc) from exc
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Export payload rejected by privacy no-leak boundary.",
        ) from exc
    _observe_human_export(
        tenant_id=auth_context.tenant_id,
        correlation_id=x_correlation_id,
        export_format=normalized_format,
        row_count=len(rows),
        artifact_bytes=(
            len(body.encode("utf-8")) if isinstance(body, str) else len(body)
        ),
    )
    return Response(
        content=body,
        media_type=media_type,
        headers={
            "X-Correlation-ID": str(x_correlation_id),
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get("/csv", operation_id="exportCSV")
async def export_csv(
    x_correlation_id: Annotated[UUID, Header(alias="X-Correlation-ID")],
    auth_context: Annotated[AuthContext, Security(get_auth_context, scopes=["viewer"])],
    db_session: Annotated[AsyncSession, Depends(get_db_session)],
    csv_schema_version: Literal[
        "b25-p11-export-csv-compat-v1", "b25-p11-export-csv-v2", "legacy-v1"
    ] = Query(default=COMPAT_CSV_SCHEMA_VERSION),
    x_attribution_session_id: Annotated[
        UUID | None, Header(alias="X-Attribution-Session-ID")
    ] = None,
):
    _assert_csv_profile_supported(csv_schema_version)
    try:
        start, end = _resolve_date_range(start_date=None, end_date=None)
        rows, _, serialized = await _run_track1_export(
            db_session=db_session,
            tenant_id=auth_context.tenant_id,
            session_scope=x_attribution_session_id,
            start=start,
            end=end,
            channels=None,
            export_format="csv",
            csv_schema_version=csv_schema_version,
        )
        if not isinstance(serialized, str):
            raise ValueError("csv_serialization_missing")
        body = serialized
    except LegacyExportLimitExceeded as exc:
        raise _limit_error(exc) from exc
    except LegacyExportDeadlineExceeded as exc:
        raise _deadline_error(exc) from exc
    _observe_human_export(
        tenant_id=auth_context.tenant_id,
        correlation_id=x_correlation_id,
        export_format="csv",
        row_count=len(rows),
        artifact_bytes=len(body.encode("utf-8")),
    )
    return Response(
        content=body,
        media_type=(
            ENRICHED_CSV_MEDIA_TYPE
            if csv_schema_version == CSV_SCHEMA_VERSION
            else COMPAT_CSV_MEDIA_TYPE
        ),
        headers={
            "X-Correlation-ID": str(x_correlation_id),
            "Content-Disposition": 'attachment; filename="skeldir-export.csv"',
        },
    )


@router.get("/json", operation_id="exportJSON")
async def export_json(
    response: Response,
    x_correlation_id: Annotated[UUID, Header(alias="X-Correlation-ID")],
    auth_context: Annotated[AuthContext, Security(get_auth_context, scopes=["viewer"])],
    db_session: Annotated[AsyncSession, Depends(get_db_session)],
    x_attribution_session_id: Annotated[
        UUID | None, Header(alias="X-Attribution-Session-ID")
    ] = None,
):
    response.headers["X-Correlation-ID"] = str(x_correlation_id)
    try:
        start, end = _resolve_date_range(start_date=None, end_date=None)
        rows, payload, _ = await _run_track1_export(
            db_session=db_session,
            tenant_id=auth_context.tenant_id,
            session_scope=x_attribution_session_id,
            start=start,
            end=end,
            channels=None,
            export_format="json",
        )
        _observe_human_export(
            tenant_id=auth_context.tenant_id,
            correlation_id=x_correlation_id,
            export_format="json",
            row_count=len(rows),
            artifact_bytes=len(str(payload).encode("utf-8")),
        )
        return payload
    except LegacyExportLimitExceeded as exc:
        raise _limit_error(exc) from exc
    except LegacyExportDeadlineExceeded as exc:
        raise _deadline_error(exc) from exc
    except ValueError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Export payload rejected by privacy no-leak boundary.",
        ) from exc


@router.get("/excel", operation_id="exportExcel")
async def export_excel(
    x_correlation_id: Annotated[UUID, Header(alias="X-Correlation-ID")],
    auth_context: Annotated[AuthContext, Security(get_auth_context, scopes=["viewer"])],
    db_session: Annotated[AsyncSession, Depends(get_db_session)],
):
    try:
        start, end = _resolve_date_range(start_date=None, end_date=None)
        rows, _, serialized = await _run_track1_export(
            db_session=db_session,
            tenant_id=auth_context.tenant_id,
            session_scope=None,
            start=start,
            end=end,
            channels=None,
            export_format="xlsx",
        )
        if not isinstance(serialized, bytes):
            raise ValueError("xlsx_serialization_missing")
        body = serialized
    except LegacyExportLimitExceeded as exc:
        raise _limit_error(exc) from exc
    except LegacyExportDeadlineExceeded as exc:
        raise _deadline_error(exc) from exc
    _observe_human_export(
        tenant_id=auth_context.tenant_id,
        correlation_id=x_correlation_id,
        export_format="xlsx",
        row_count=len(rows),
        artifact_bytes=len(body),
    )
    return Response(
        content=body,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "X-Correlation-ID": str(x_correlation_id),
            "Content-Disposition": 'attachment; filename="skeldir-export.xlsx"',
        },
    )
